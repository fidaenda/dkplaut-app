import streamlit as st
import pandas as pd
import tempfile
import os
from io import BytesIO
from openpyxl import load_workbook

st.title("📊 Olah Data Perikanan Budidaya Laut Jawa Timur")

# File Uploader
uploaded_files = st.file_uploader("📂 Upload satu atau lebih file Excel", type=["xlsx", "xls"], accept_multiple_files=True)
template_file = st.file_uploader("📂 Upload file template", type=["xlsx"])  # Hanya terima format xlsx

def extract_data(laut):
    """Fungsi untuk mengekstrak data dari sheet '1. LAUT'."""
    Y = laut.iloc[1, 0]
    X1 = laut.iloc[51, 2] if 51 < len(laut) and isinstance(laut.iloc[51, 2], (int, float)) else 0
    X2 = laut.iloc[4, 2] if 4 < len(laut) and isinstance(laut.iloc[4, 2], (int, float)) else 0
    X3 = laut.iloc[5, 2] if 5 < len(laut) and isinstance(laut.iloc[5, 2], (int, float)) else 0
    
    # Definisi X4-X13
    X4 = laut.iloc[9:27, 1:4].drop(index=21).dropna(how='all').set_axis(["Jenis Ikan", "Volume (Ton)", "Nilai (Rp. 000)"], axis=1).astype({"Jenis Ikan": str}).assign(**{"Volume (Ton)": lambda df: pd.to_numeric(df["Volume (Ton)"], errors='coerce').fillna(0).round(2), "Nilai (Rp. 000)": lambda df: pd.to_numeric(df["Nilai (Rp. 000)"], errors='coerce').fillna(0).astype("int64")})
    X5 = laut.iloc[33:50, 1:3].drop(index=45).dropna(how='all').set_axis(["Jenis Ikan", "Volume (x1000 Ekor)"], axis=1).astype({"Jenis Ikan": str}).assign(**{"Volume (x1000 Ekor)": lambda df: pd.to_numeric(df["Volume (x1000 Ekor)"], errors='coerce').fillna(0).round(2)})
    X6 = laut.iloc[54:58, 1:4].set_axis(["Besarnya Usaha (RTP)", "KJA Laut", "Budidaya Rula"], axis=1).astype({"Besarnya Usaha (RTP)": str}).assign(**{"KJA Laut": lambda df: pd.to_numeric(df["KJA Laut"], errors='coerce').fillna(0).astype(int), "Budidaya Rula": lambda df: pd.to_numeric(df["Budidaya Rula"], errors='coerce').fillna(0).astype(int)})
    X7 = laut.iloc[60:64, 1:4].set_axis(["Jenis Yang di tanam (RTP)", "KJA Laut", "Budidaya Rula"], axis=1).astype({"Jenis Yang di tanam (RTP)": str}).assign(**{"KJA Laut": lambda df: pd.to_numeric(df["KJA Laut"], errors='coerce').fillna(0).astype(int), "Budidaya Rula": lambda df: pd.to_numeric(df["Budidaya Rula"], errors='coerce').fillna(0).astype(int)})
    X8 = laut.iloc[66:68, 1:4].set_axis(["Penggunaan Obat (RTP)", "KJA Laut", "Budidaya Rula"], axis=1).astype({"Penggunaan Obat (RTP)": str}).assign(**{"KJA Laut": lambda df: pd.to_numeric(df["KJA Laut"], errors='coerce').fillna(0).astype(int), "Budidaya Rula": lambda df: pd.to_numeric(df["Budidaya Rula"], errors='coerce').fillna(0).astype(int)})
    X9 = laut.iloc[69:70, 1:4].set_axis(["Jumlah Obat (kg/L)*", "KJA Laut", "Budidaya Rula"], axis=1).astype({"Jumlah Obat (kg/L)*": str}).assign(**{"KJA Laut": lambda df: pd.to_numeric(df["KJA Laut"], errors='coerce').fillna(0).astype(int), "Budidaya Rula": lambda df: pd.to_numeric(df["Budidaya Rula"], errors='coerce').fillna(0).astype(int)})
    X10 = laut.iloc[73:75, 1:4].set_axis(["Laut", "KJA Laut", "Budidaya Rula"], axis=1).astype({"Laut": str}).assign(**{"KJA Laut": lambda df: pd.to_numeric(df["KJA Laut"], errors='coerce').fillna(0).astype(int), "Budidaya Rula": lambda df: pd.to_numeric(df["Budidaya Rula"], errors='coerce').fillna(0).astype(int)})
    X11 = laut.iloc[79:81, 1:4].set_axis(["Budidaya Laut", "KJA Laut", "Budidaya Rula"], axis=1).astype({"Budidaya Laut": str}).assign(**{"KJA Laut": lambda df: pd.to_numeric(df["KJA Laut"], errors='coerce').fillna(0).round(2), "Budidaya Rula": lambda df: pd.to_numeric(df["Budidaya Rula"], errors='coerce').fillna(0).round(2)})
    X12 = laut.iloc[85:90, 1:4].set_axis(["Pakan yang Digunakan (kg)", "KJA", "Rumput Laut"], axis=1).astype({"Pakan yang Digunakan (kg)": str}).assign(**{"KJA": lambda df: pd.to_numeric(df["KJA"], errors='coerce').fillna(0).round(2), "Rumput Laut": lambda df: pd.to_numeric(df["Rumput Laut"], errors='coerce').fillna(0).round(2)})
    X13 = laut.iloc[91:98, 1:4].set_axis(["Penggunaan Pestisida dan obat-obatan yang digunakan", "KJA", "Rumput Laut"], axis=1).astype({"Penggunaan Pestisida dan obat-obatan yang digunakan": str}).assign(**{"KJA": lambda df: pd.to_numeric(df["KJA"], errors='coerce').fillna(0).round(2), "Rumput Laut": lambda df: pd.to_numeric(df["Rumput Laut"], errors='coerce').fillna(0).round(2)})

    return Y, X1, X2, X3, X4, X5, X6, X7, X8, X9, X10, X11, X12, X13

if template_file:
    temp_dir = tempfile.mkdtemp()
    template_path = os.path.join(temp_dir, "template.xlsx")
    with open(template_path, "wb") as f:
        f.write(template_file.getbuffer())
    template_wb = load_workbook(template_path)
    A = template_wb["7.1"]
    B = template_wb["7.2"]
    C = template_wb["7.3"]
    D = template_wb["7.5"]

    if uploaded_files:
        file_paths = {uploaded_file.name: uploaded_file for uploaded_file in uploaded_files}
        selected_file = st.selectbox("📂 Pilih file untuk ditampilkan", list(file_paths.keys()))

        if selected_file:
            st.write(f"Menampilkan data dari: {selected_file}")
            xls = pd.ExcelFile(file_paths[selected_file])
            sheet_names = xls.sheet_names
            
            for selected_sheet in sheet_names:
                if selected_sheet.strip() == "1. LAUT":
                    try:
                        laut = pd.read_excel(xls, sheet_name=selected_sheet)
                        Y, X1, X2, X3, X4, X5, X6, X7, X8, X9, X10, X11, X12, X13 = extract_data(laut)

                        # Definisi tabel
                        tables = {
                            "Kabupaten/Kota": Y,
                            "Jumlah RTP Budidaya Laut": X1,
                            "Potensi Lahan Budidaya Laut": X2,
                            "Potensi Produksi Budidaya Laut": X3,
                            "Produksi dan Nilai Produksi": X4,
                            "Benih Ikan yang Ditanam": X5,
                            "Jumlah RTP Berdasarkan Besar Usaha": X6,
                            "Jumlah RTP Berdasarkan Jenis yang Ditanam": X7,
                            "Jumlah RTP Berdasarkan Penggunaan Obat": X8,
                            "Jumlah Obat yang Digunakan": X9,
                            "Jumlah Petani Budidaya Laut": X10,
                            "Luas Areal Budidaya Laut": X11,
                            "Penggunaan Pakan di Budidaya Laut": X12,
                            "Penggunaan Obat di Budidaya Laut": X13
                        }

                        selected_variable = st.selectbox("📊 Pilih Tabel Data", list(tables.keys()), key=selected_file)
                        st.subheader(f"📌 Data {selected_variable} - {Y}")
                        st.write(tables[selected_variable])
                    except Exception as e:
                        st.error(f"❌ Error saat memproses data dari {selected_file} pada sheet {selected_sheet}: {e}")
    
        # Konfirmasi sebelum update
        if st.button("Update Tabel"):
            for file_name, uploaded_file in file_paths.items():
                xls = pd.ExcelFile(uploaded_file)
                for selected_sheet in xls.sheet_names:
                    if selected_sheet.strip() == "1. LAUT":
                        laut = pd.read_excel(xls, sheet_name=selected_sheet)
                        Y, X1, X2, X3, X4, X5, X6, X7, X8, X9, X10, X11, X12, X13 = extract_data(laut)

                        for row in A.iter_rows(min_row=2, max_row=A.max_row):
                            if row[0].value and str(row[0].value).strip().lower() == str(Y).strip().lower():
                                row[1].value = X1  # Update RTP
                                row[2].value = X2  # Update Potensi Lahan
                                row[3].value = X5.iloc[2, 1]  # Update Benih - Kerapu
                                row[4].value = X5.iloc[3, 1]  # Update Benih - Kakap
                                row[5].value = X5.iloc[9, 1]  # Update Benih - Bawal Bintang
                                row[6].value = X5.iloc[0, 1]  # Update Benih - Bandeng
                                row[7].value = X5.iloc[10, 1] + X5.iloc[12, 1] + X5.iloc[13, 1]  # Update Benih - Rumput Laut
                                row[8].value = X5.iloc[6, 1]  # Update Benih - Udang barong
                                row[9].value = X5.iloc[7, 1]  # Update Benih - Udang Lainnya
                                row[10].value = 0  # Update Benih - Kerang Darah (gaada di list benih?)
                                row[11].value = 0  # Update Benih - Kerang Hijau (gaada di list benih?)
                                row[12].value = X5.iloc[1, 1] + X5.iloc[4, 1] + X5.iloc[5, 1] + X5.iloc[8, 1] + X5.iloc[11, 1] + X5.iloc[14, 1]  # Update Benih - Lainnya
                                break

                        for row in B.iter_rows(min_row=2, max_row=B.max_row):
                            if row[0].value and str(row[0].value).strip().lower() == str(Y).strip().lower():
                                row[1].value = X12.iloc[0, 1] + X12.iloc[0, 2] # Update Pakan - Ikan Rucah
                                row[2].value = X12.iloc[1, 1] + X12.iloc[1, 2] # Update Pakan - Pelet
                                row[3].value = X12.iloc[2, 1] + X12.iloc[2, 2] # Update Pakan - Dedak
                                row[4].value = X12.iloc[3, 1] + X12.iloc[3, 2] # Update Pakan - Lainnya
                                row[5].value = X12.iloc[4, 1] + X12.iloc[4, 2] # Update Pakan - Probiotik
                                row[6].value = 0 # Update Kapur (gaada di list?)
                                row[7].value = X13.iloc[0, 1] + X13.iloc[0, 2] # Update Obat - Pestisida
                                row[8].value = X13.iloc[1, 1] + X13.iloc[1, 2] # Update Obat - Desinfektan
                                row[9].value = X13.iloc[2, 1] + X13.iloc[2, 2] # Update Obat - Obat
                                row[10].value = X13.iloc[3, 1] + X13.iloc[3, 2] # Update Obat - Hormon
                                row[11].value = X13.iloc[5, 1] + X13.iloc[5, 2] # Update Obat - Listrik(KVA) >>> KEKNYA INI PENYEBAB ERROR
                                row[12].value = X13.iloc[6, 1] + X13.iloc[6, 2] # Update Obat - BBM
                                break

                        for row in C.iter_rows(min_row=2, max_row=C.max_row):
                            if row[0].value and str(row[0].value).strip().lower() == str(Y).strip().lower():
                                row[2].value = X4.iloc[0, 1]  # Update Produksi - Bandeng
                                row[3].value = X4.iloc[1, 1]  # Update Produksi - Belanak
                                row[4].value = X4.iloc[2, 1]  # Update Produksi - Kerapu (Tikus)
                                row[5].value = X4.iloc[3, 1]  # Update Produksi - Kakap
                                row[6].value = X4.iloc[4, 1]  # Update Produksi - Baronang
                                row[7].value = X4.iloc[5, 1] + X4.iloc[9, 1] # Update Produksi - Ikan Lainnya + Bawal Bintang
                                row[8].value = X4.iloc[6, 1]  # Update Produksi - Udang Barong
                                row[9].value = X4.iloc[7, 1]  # Update Produksi - Udang Lainnya
                                row[10].value = X4.iloc[8, 1]  # Update Produksi - Rajungan
                                row[11].value = 0  # Update Produksi - Kerang Darah (gaada di list produksi?)
                                row[12].value = X4.iloc[14, 1]  # Update Produksi - Kerang Hijau
                                row[13].value = 0  # Update Produksi - Remis (gaada di list produksi?)
                                row[14].value = 0  # Update Produksi - Mutiara (gaada di list produksi?)
                                row[15].value = X4.iloc[11, 1]  # Update Produksi - Lainnya (Kepiting)
                                row[16].value = X4.iloc[10, 1] + X4.iloc[12, 1] + X4.iloc[13, 1] # Update Produksi - Rumput Laut
                                break

                        for row in D.iter_rows(min_row=2, max_row=D.max_row):
                            if row[0].value and str(row[0].value).strip().lower() == str(Y).strip().lower():
                                row[2].value = X4.iloc[0, 2]  # Update Nilai Produksi - Bandeng
                                row[3].value = X4.iloc[1, 2]  # Update Nilai Produksi - Belanak
                                row[4].value = X4.iloc[2, 2]  # Update Nilai Produksi - Kerapu
                                row[5].value = X4.iloc[3, 2]  # Update Nilai Produksi - Kakap
                                row[6].value = X4.iloc[4, 2]  # Update Nilai Produksi - Baronang
                                row[7].value = X4.iloc[5, 2] + X4.iloc[9, 2] # Update Nilai Produksi - Ikan Lainnya + Bawal Bintang
                                row[8].value = X4.iloc[6, 2]  # Update Nilai Produksi - Udang Barong
                                row[9].value = X4.iloc[7, 2]  # Update Nilai Produksi - Udang Lainya
                                row[10].value = X4.iloc[8, 2]  # Update Nilai Produksi - Rajungan
                                row[11].value = 0  # Update Nilai Produksi - Kerang Darah (gaada di list produksi?)
                                row[12].value = X4.iloc[14, 2]  # Update Nilai Produksi - Kerang Hijau
                                row[13].value = 0  # Update Nilai Produksi - Remis (gaada di list produksi?)
                                row[14].value = 0  # Update Nilai Produksi - Mutiara (gaada di list produksi?)
                                row[15].value = X4.iloc[11, 2]  # Update Nilai Produksi - Lainnya (Kepiting)
                                row[16].value = X4.iloc[10, 2] + X4.iloc[12, 2] + X4.iloc[13, 2] # Update Nilai Produksi - Rumput Laut
                                break

        # Simpan hasil
        output = BytesIO()
        template_wb.save(output)
        output.seek(0)

        st.download_button(
            label="📥 Download File Hasil",
            data=output.getvalue(),
            file_name="Olah Data Perikanan Budidaya Laut Jawa Timur.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
